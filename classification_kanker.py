import pandas as pd
import numpy as np

import math

from openpyxl import load_workbook
workbook = load_workbook(filename="datauji.xlsx")

sheet = workbook.active

from sklearn.model_selection import train_test_split

# Get neccesary columns
datarule=pd.read_excel('rulesNaive.xls', sheet_name = 'Sheet1')

arr_datarule = datarule.to_numpy()

def calculate_numeric(x, mean, standar):
    return 1 / (math.sqrt( math.pi * 2 * standar )  ) * math.exp(-( pow((x-mean), 2)/ (2 * pow(standar, 2)) ))

def getmean_numeric(rules, header):
    result = {
      "living": 0.0,
      "died": 0.0
    }
    if header == "tumor_size":
        result["living"]= rules[0][2]
        result["died"]= rules[0][3]
    elif header == "tumor_stage":
        result["living"]= rules[2][2]
        result["died"]= rules[2][3]
    elif header == "mutation_count":
        result["living"]= rules[18][2]
        result["died"]= rules[18][3]
    elif header == "neoplasm_histologic_grade":
        result["living"]= rules[34][2]
        result["died"]= rules[34][3]
    elif header == "age_at_diagnosis":
        result["living"]= rules[65][2]
        result["died"]= rules[65][3]
    elif header == "cohort":
        result["living"]= rules[78][2]
        result["died"]= rules[78][3]
    elif header == "lymph_nodes_examined_positive":
        result["living"]= rules[106][2]
        result["died"]= rules[106][3]
    elif header == "nottingham_prognostic_index":
        result["living"]= rules[108][2]
        result["died"]= rules[108][3]
    return result

def getstandar_numeric(rules, header):
    result = {
      "living": 0.0,
      "died": 0.0
    }
    if header == "tumor_size":
        result["living"]= rules[1][2]
        result["died"]= rules[1][3]
    elif header == "tumor_stage":
        result["living"]= rules[3][2]
        result["died"]= rules[3][3]
    elif header == "mutation_count":
        result["living"]= rules[19][2]
        result["died"]= rules[19][3]
    elif header == "neoplasm_histologic_grade":
        result["living"]= rules[35][2]
        result["died"]= rules[35][3]
    elif header == "age_at_diagnosis":
        result["living"]= rules[66][2]
        result["died"]= rules[66][3]
    elif header == "cohort":
        result["living"]= rules[79][2]
        result["died"]= rules[79][3]
    elif header == "lymph_nodes_examined_positive":
        result["living"]= rules[107][2]
        result["died"]= rules[107][3]
    elif header == "nottingham_prognostic_index":
        result["living"]= rules[109][2]
        result["died"]= rules[109][3]
    return result

def getProbs_nominal(rules, header, value):
    result = {
      "living": 0.0,
      "died": 0.0
    }
    condition = "value={}".format(value)
    if header not in {"tumor_size", "tumor_stage", "mutation_count", "neoplasm_histologic_grade", "age_at_diagnosis", "cohort",
    "lymph_nodes_examined_positive", "nottingham_prognostic_index"}:
        for x in rules:
            if x[0] == header and x[1] == condition:
                result["living"] = x[2]
                result["died"] = x[3]

    return result

tophead = ()
for value in sheet.iter_rows(max_row=1,
                              min_col=1,
                              max_col=29,
                              values_only=True):
   tophead = value

row = 1
disscorrect = 0
for value in sheet.iter_rows(min_row=2,
                              min_col=1,
                              max_col=29,
                              values_only=True):
   row = row + 1
   probs_living = float(1)
   probs_died = float(1)
   for idx in range(len(value)):
       if idx in {0, 7, 10, 18, 19, 20, 25, 26}:
           mean = getmean_numeric(arr_datarule, tophead[idx])
           standar = getstandar_numeric(arr_datarule, tophead[idx])
           if value[idx] is None:
               probs_living = probs_living * calculate_numeric(mean['living'], mean['living'], standar['living'])
               probs_died = probs_died * calculate_numeric(mean['died'], mean['died'], standar['died'])
           else:
               valnumeric = value[idx]
               probs_living = probs_living * calculate_numeric(valnumeric, mean['living'], standar['living'])
               probs_died = probs_died * calculate_numeric(valnumeric, mean['died'], standar['died'])
       elif idx in {1,2,3,4,5,6,8,9,11,12,13,14,15,16,17,21,22,23,24}:
           if value[idx] is None:
               valnominal = "UNDEF"
           elif value[idx] == 0:
               valnominal = "false"
           elif value[idx] == 1:
               valnominal = "true"
           else:
               valnominal = value[idx]
           probablilitas_evidence = getProbs_nominal(arr_datarule, tophead[idx], valnominal)
           probs_living = probs_living * probablilitas_evidence["living"]
           probs_died = probs_died * probablilitas_evidence["died"]
   sheet['AC{}'.format(row)] = probs_living
   sheet['AD{}'.format(row)] = probs_died
   classpred = None
   if probs_living > probs_died:
       classpred = "Living"
   else:
       classpred = "Died of Disease"
   sheet['AE{}'.format(row)] = classpred


workbook.save("datauji.xlsx")
